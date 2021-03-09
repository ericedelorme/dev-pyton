#!/usr/bin/env python3


import os
import os.path
import re
import argparse
import datetime
from sys import stdout
from os import stat
from time import sleep
from datetime import datetime
from PIL import Image
from PIL.ExifTags import TAGS
from hachoir.parser import createParser
from hachoir.metadata import extractMetadata
import filecmp
import hashlib
import pathlib
from openpyxl import Workbook

default_tiff_datetime_tag_value = datetime(1970, 1, 1)


def argparse_is_dir(dirname):
    """Checks if a path is an actual directory"""
    if not os.path.isdir(dirname):
        msg = "{0} is not a directory".format(dirname)
        raise argparse.ArgumentTypeError(msg)
    else:
        return dirname


def sys_stdout_progress_bar(current_value, max_value, unit="", text=""):
    # print("sys_stdout_progress_bar(" + str(current_value) + "," + str(max_value)  + "," + unit + ")")
    bar_length = 72  # Modify this to change the length of the progress bar

    if not isinstance(current_value, int):
        current_value = int(current_value)
    if not isinstance(max_value, int):
        max_value = int(max_value)

    if current_value < 0:
        status = "error: current_value var must be >0\r\n"
        percent = 0
    elif current_value >= max_value:
        status = "(" + str(current_value) + " " + unit + " / " + str(max_value) + " " + unit + ")" + text
        percent = 1
    else:
        status = "(" + str(current_value) + " " + unit + " / " + str(max_value) + " " + unit + ")" + text
        percent = current_value / max_value

    block = int(round(bar_length * percent))
    text = "\r[{0}] {1}% {2}".format("=" * block + " " * (bar_length - block), int(percent * 100), status)
    stdout.write(text)
    stdout.flush()


def is_video_file(file_path):
    # return file_path.lower().endswith(('.mp4', 'm4v', 'mov'))
    return file_path.lower().endswith('.mp4')


def is_image_file(file_path):
    return file_path.lower().endswith(('.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif', '.heic'))


def get_datetime_tag_value(file_path):
    datetime_tag_value: datetime = default_tiff_datetime_tag_value
    datetimeoriginal_tag_value: datetime = default_tiff_datetime_tag_value
    datetimedigitized_tag_value: datetime = default_tiff_datetime_tag_value
    earliest_tag_value: datetime = default_tiff_datetime_tag_value
    video_creation_date_value = default_tiff_datetime_tag_value
    # retrieve TIFF 'DateTime', 'DateTimeOriginal' and  'DateTimeDigitized' tag value of file_path and convert as date_time object
    if is_image_file(file_path):
        try:
            for (tag, value) in Image.open(file_path)._getexif().items():
                tag_name = str(TAGS.get(tag))
                str_value = ""
                try:
                    if tag_name == 'DateTime' and value != '':
                        # some bug with midnight hours  00:15:07 appear like 24:15;07, so ' 24:' is replace by ' 00:'
                        str_value = str(value).replace(' 24:', ' 00:')
                        # some bug extra char at the end of string
                        if len(str_value) > 19:
                            str_value = str_value[0:18]
                        datetime_tag_value = datetime.strptime(str_value, '%Y:%m:%d %H:%M:%S')
                        if earliest_tag_value == default_tiff_datetime_tag_value or datetime_tag_value < earliest_tag_value:
                            earliest_tag_value = datetime_tag_value
                    elif tag_name == 'DateTimeOriginal' and value != '':
                        # some bug with midnight hours  00:15:07 appear like 24:15;07, so ' 24:' is replace by ' 00:'
                        str_value = str(value).strip(' ').replace(' 24:', ' 00:')
                        if len(str_value) > 19:
                            str_value = str_value[0:18]
                        datetimeoriginal_tag_value = datetime.strptime(str_value, '%Y:%m:%d %H:%M:%S')
                        if earliest_tag_value == default_tiff_datetime_tag_value or datetimeoriginal_tag_value < earliest_tag_value:
                            earliest_tag_value = datetimeoriginal_tag_value
                    elif tag_name == 'DateTimeDigitized' and value != '':
                        # some bug with midnight hours  00:15:07 appear like 24:15;07, so ' 24:' is replace by ' 00:'
                        str_value = str(value).strip(' ').replace(' 24:', ' 00:')
                        if len(str_value) > 19:
                            str_value = str_value[0:18]
                        datetimedigitized_tag_value = datetime.strptime(str_value, '%Y:%m:%d %H:%M:%S')
                        if earliest_tag_value == default_tiff_datetime_tag_value or datetimedigitized_tag_value < earliest_tag_value:
                            earliest_tag_value = datetimedigitized_tag_value
                except ValueError:
                    print("get_datetime_tag_value('" + file_path + "') : ValueError datetime.strptime(tag " + tag_name + "='" + str_value + "', '%Y:%m:%d %H:%M:%S')")
            return earliest_tag_value, datetime_tag_value, datetimeoriginal_tag_value, datetimedigitized_tag_value
        except IOError:
            # file_path not an image file
            # print("IOError in get_datetime_tag_value('"+file_path+"'")
            return earliest_tag_value, datetime_tag_value, datetimeoriginal_tag_value, datetimedigitized_tag_value
        except AttributeError:
            # print(file_path + " : AttributeError")
            # print("AttributeError in get_datetime_tag_value('"+file_path+"'")
            return earliest_tag_value, datetime_tag_value, datetimeoriginal_tag_value, datetimedigitized_tag_value
    elif is_video_file(file_path):
        try:
            parser = createParser(file_path)
            metadata = extractMetadata(parser)
            for line in metadata.exportPlaintext():
                # Creation date tag is like "- Creation date: 2018-01-04 15:26:52"
                if "Creation date: " in line:
                    str_value = line.split("Creation date: ")[1]
                    # some bug with midnight hours  00:15:07 appear like 24:15;07, so ' 24:' is replace by ' 00:'
                    str_value = str_value.replace(' 24:', ' 00:')
                    # some bug extra char at the end of string
                    if len(str_value) > 19:
                        str_value = str_value[0:18]
                    try:
                        video_creation_date_value = datetime.strptime(str_value, '%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        print("get_datetime_tag_value('" + file_path + "') : ValueError datetime.strptime(" + str_value + "', '%Y-%m-%d %H:%M:%S')")
                    break
        except Exception as err:
            print("get_datetime_tag_value('" + file_path + "') : Metadata extraction error : " + str(err), file=stderr)
        return video_creation_date_value, default_tiff_datetime_tag_value, default_tiff_datetime_tag_value, default_tiff_datetime_tag_value


def get_tiff_tag_value(file_path, tag_name):
    if is_image_file(file_path):
        try:
            for (tag, value) in Image.open(file_path)._getexif().items():
                if str(TAGS.get(tag)) == tag_name:
                    return str(value)
            return "TIFF tag " + tag_name + " Not Found"
        except IOError:
            # file_path not an image file
            return ""
        except AttributeError:
            # print(file_path + " : AttributeError")
            return ""
    return ""


def get_list_of_all_images_or_videos_with_date_and_size(root_folder_path_list, find_images, find_videos, output_excel_file_path, output_csv_file_path):
    # print('get_list_of_all_images_or_videos_with_date_and_size(' + root_folder_path + ')\n')
    all_path_are_valid = True
    concatenated_root_folder_path = ""
    for root_folder_path in root_folder_path_list:
        if concatenated_root_folder_path == "":
            concatenated_root_folder_path = "'" + root_folder_path + "'"
        else:
            concatenated_root_folder_path = concatenated_root_folder_path + ", '" + root_folder_path + "'"

        if not os.path.isdir(root_folder_path):
            all_path_are_valid = False


    if all_path_are_valid:
        nb_files: int = 0
        files_list = []
        files_list_with_data = []
        stdout.write("\n")
        stdout.flush()
        stdout.write("\rAnalysis of " + concatenated_root_folder_path + " : in progress...")
        stdout.flush()
        for root_folder_path in root_folder_path_list:
            folder_walk = os.walk(root_folder_path)
            for (d_folder, d_dirs, d_files) in folder_walk:
                for d_file in d_files:
                    if 'synoreport' not in d_folder and '#recycle' not in d_folder:
                        if (find_images == True and is_image_file(d_file) == True) or (find_videos == True and is_video_file(d_file) == True):
                            nb_files = nb_files + 1
                            # appende full_file_path, folder_path, root_folder_path, first_relative_folder_name, file_name
                            relative_root_root_folder_path = os.path.relpath(d_folder, root_folder_path)
                            first_relative_folder_name = relative_root_root_folder_path.split(os.sep)[0]
                            files_list.append((os.path.join(d_folder, d_file), relative_root_root_folder_path, first_relative_folder_name, d_file))

        stdout.write("\rAnalysis of " + concatenated_root_folder_path + " : " + str(nb_files) + " files found, retrieving date, size in progress...\n")

        step_progress = int(round(nb_files / 500))
        if step_progress == 0:
            step_progress = 1
        prev_nb_files = 0
        curr_nb_files = 0
        index_duplictate = 0
        total_file_size_mb = 0.0
        total_file_size_gb = 0.0
        prev_date_time = datetime.now()
        sys_stdout_progress_bar(curr_nb_files, nb_files, "files", " : " + str(index_duplictate) + " duplicate(s) found")
        for d_file_path, d_relative_root_root_folder_path, d_first_relative_folder_name, d_file in files_list:
            curr_nb_files = curr_nb_files + 1

            os_stat = os.stat(d_file_path)
            file_last_modified = datetime.fromtimestamp(os_stat.st_mtime)
            file_size = os_stat.st_size
            file_size_mb = float(file_size) / float(1000000)
            total_file_size_mb = total_file_size_mb + file_size_mb
            total_file_size_gb = total_file_size_mb / 1000
            tiff_datetime = get_datetime_tag_value(d_file_path)
            sha256_sum = hashlib.sha256(pathlib.Path(d_file_path).read_bytes()).hexdigest()

            # does file with same size and same sha256 exists
            id_duplicate = 0
            for i in range(len(files_list_with_data)):
                i_file_size = files_list_with_data[i][5]
                i_sha256_sum = files_list_with_data[i][7]
                i_id_duplicate = files_list_with_data[i][8]
                if i_file_size == file_size and i_sha256_sum == sha256_sum:
                    if i_id_duplicate == 0:
                        index_duplictate = index_duplictate + 1
                        id_duplicate = index_duplictate
                        files_list_with_data[i] = (
                        files_list_with_data[i][0], files_list_with_data[i][1], files_list_with_data[i][2],
                        files_list_with_data[i][3], files_list_with_data[i][4],
                        files_list_with_data[i][5], files_list_with_data[i][6], files_list_with_data[i][7],
                        id_duplicate,
                        files_list_with_data[i][9], files_list_with_data[i][10], files_list_with_data[i][11],
                        files_list_with_data[i][12], files_list_with_data[i][13])
                    else:
                        id_duplicate = i_id_duplicate
                    break

            files_list_with_data.append(
                (d_file_path, root_folder_path, d_relative_root_root_folder_path, d_first_relative_folder_name, d_file,
                 file_size, file_size_mb, sha256_sum, id_duplicate,
                 file_last_modified, tiff_datetime[0], tiff_datetime[1], tiff_datetime[2], tiff_datetime[3]))
            curr_date_time=datetime.now()
            if ((prev_nb_files + step_progress) <= curr_nb_files) or (1.0 <= (curr_date_time-prev_date_time).total_seconds()):
                prev_nb_files = curr_nb_files
                prev_date_time = curr_date_time
                sys_stdout_progress_bar(curr_nb_files, nb_files, "files",
                                        " : " + str(index_duplictate) + " duplicate(s) found")
        if total_file_size_mb < 999.0:
            str_total_file_size = str(round(total_file_size_mb, 0)) + " Mo"
        else:
            str_total_file_size = str(round(total_file_size_gb, 3)) + " Go"
        sys_stdout_progress_bar(curr_nb_files, nb_files, "files",
                                " : " + str_total_file_size + " : " + str(index_duplictate) + " duplicate(s) found")
        stdout.write("\n")

        first_line = "file path;root path;relative folder path;name of first folder;file name;file size (Octet);file size (Mo);sha256;id duplicate;date last modified;Earliest DateTime from TIFF Tag value (images)/Creation date (videos);DateTime TIFF Tag value;DateTimeOriginal TIFF Tag value;DateTimeDigitized TIFF Tag value"
        split_first_line = first_line.split(';')
        nb_columns = len(split_first_line)
        file_path_column = 0
        if output_excel_file_path != '':
            wb = Workbook()
            ws = wb.active
            for j in range(nb_columns):
                ws.cell(1, j + 1).value = split_first_line[j]
                if split_first_line[j] == "file path":
                    file_path_column = j + 1
            is_excel_output = True
        else:
            is_excel_output = False

        if output_csv_file_path != '':
            output_csv_file = open(output_csv_file_path, 'w')
            output_csv_file.write(first_line + "\n")
            is_csv_output = True
        else:
            is_csv_output = False

        # for each file's tuple of data (path, size,date,...) in list
        for i in range(nb_files):
            if is_excel_output:
                for j in range(nb_columns):
                    ws.cell(2 + i, j + 1).value = files_list_with_data[i][j]
                    # secial case for file_path column, add hyperlink to file
                    if (j + 1) == file_path_column:
                        ws.cell(2 + i, 1 + j).hyperlink = files_list_with_data[i][j]
                        ws.cell(2 + i, 1 + j).style = "Hyperlink"

            if is_csv_output:
                line = ''
                for j in range(nb_columns):
                    if j == 0:
                        line = str(files_list_with_data[i][j])
                    else:
                        line = line + ';' + str(files_list_with_data[i][j])
                output_csv_file.write(line + '\n')
        if is_excel_output:
            wb.save(output_excel_file_path)
            stdout.write("File '" + output_excel_file_path + "' Generated\n")
        if is_csv_output:
            output_csv_file.close()
            stdout.write("File '" + output_csv_file_path + "' Generated\n")



def main():
    # ArgumentParser(prog=None, usage=None, description=None, epilog=None, parents=[], formatter_class=argparse.HelpFormatter, prefix_chars='-', fromfile_prefix_chars=None, argument_default=None, conflict_handler='error', add_help=True, allow_abbrev=True, exit_on_error=True)
    parser = argparse.ArgumentParser()

    # parser.add_argument('-f', '--folder_path', default=os.getcwd(), help='specify the path of the folder in which to find duplicate files. Default is current dir')
    parser.add_argument('dirname', nargs='*', type=argparse_is_dir, default=[os.getcwd()], help='specify the path of the folder in which to find duplicate files. Default is current dir')
    group = parser.add_mutually_exclusive_group()
    parser.add_argument('-e', '--excel_output', help='specify the Excel output .xlsx file path')
    parser.add_argument('-c', '--csv_output', help='specify the CSV output .csv file path')
    group.add_argument('-i', '--images_only', action='store_true', help='find only images')
    group.add_argument('-v', '--videos_only', action='store_true', help='find only videos')
    args = parser.parse_args()

    if ((args.csv_output != '' and args.csv_output is not None)
            or (args.excel_output != '' and args.excel_output is not None)):
        if args.excel_output != '' and args.excel_output is not None:
            args.excel_output = os.path.realpath(args.excel_output)
        else:
            args.excel_output = ''

        if args.csv_output != '' and args.csv_output is not None:
            args.csv_output = os.path.realpath(args.csv_output)
        else:
            args.csv_output = ''

        get_list_of_all_images_or_videos_with_date_and_size([os.path.realpath(element) for element in args.dirname], not args.videos_only, not args.images_only, args.excel_output, args.csv_output)
    else:
        parser.print_help()


if __name__ == "__main__":
    # execute only if run as a script
    main()
