#!/usr/bin/env python3

import os.path
import re
import argparse

from sys import stdout
from os import stat
from time import sleep
from datetime import datetime

from PIL import Image
from PIL.ExifTags import TAGS
from openpyxl import Workbook

default_tiff_datetime_tag_value = datetime(1970, 1, 1)


def sys_stdout_progress_bar(current_value, max_value, unit="", text = ""):
    # print("sys_stdout_progress_bar(" + str(current_value) + "," + str(max_value)  + "," + unit + ")")
    bar_length = 72  # Modify this to change the length of the progress bar

    if not isinstance(current_value, int):
        current_value = int(current_value)
    if not isinstance(max_value, int):
        max_value = int(max_value)

    if current_value < 0:
        status = "error: current_value var must be >0\r\n"
        percent = 0
    elif current_value >= max_value:
        status = "(" + str(current_value) + " " + unit + " / " + str(max_value) + " " + unit + ")" + text
        percent = 1
    else:
        status = "(" + str(current_value) + " " + unit + " / " + str(max_value) + " " + unit + ")" + text
        percent = current_value / max_value

    block = int(round(bar_length * percent))
    text = "\r[{0}] {1}% {2}".format("=" * block + " " * (bar_length - block), int(percent * 100), status)
    stdout.write(text)
    stdout.flush()


def is_image_file(file_path):
    return file_path.lower().endswith(('.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif', '.heic'))


def get_tiff_datetime_datetimeoriginal_datetimedigitized(file_path):
    datetime_tag_value: datetime = default_tiff_datetime_tag_value
    datetimeoriginal_tag_value: datetime = default_tiff_datetime_tag_value
    datetimedigitized_tag_value: datetime = default_tiff_datetime_tag_value
    earliest_tag_value: datetime = default_tiff_datetime_tag_value
    # retrieve TIFF 'DateTime', 'DateTimeOriginal' and  'DateTimeDigitized' tag value of file_path and convert as date_time object
    if is_image_file(file_path):
        try:
            for (tag, value) in Image.open(file_path)._getexif().items():
                tag_name = str(TAGS.get(tag))
                try:
                    if tag_name == 'DateTime' and value != '':
                        # some bug with midnight hours  00:15:07 appear like 24:15;07, so ' 24:' is replace by ' 00:'
                        str_value=str(value).replace(' 24:', ' 00:')
                        # some bug extra char at the end of string
                        if len(str_value)>19:
                            str_value=str_value[0:18]
                        datetime_tag_value = datetime.strptime(str_value, '%Y:%m:%d %H:%M:%S')
                        if earliest_tag_value == default_tiff_datetime_tag_value or datetime_tag_value < earliest_tag_value:
                            earliest_tag_value = datetime_tag_value
                    elif tag_name == 'DateTimeOriginal' and value != '':
                        # some bug with midnight hours  00:15:07 appear like 24:15;07, so ' 24:' is replace by ' 00:'
                        str_value=str(value).strip(' ').replace(' 24:', ' 00:')
                        if len(str_value)>19:
                            str_value=str_value[0:18]
                        datetimeoriginal_tag_value = datetime.strptime(str_value, '%Y:%m:%d %H:%M:%S')
                        if earliest_tag_value == default_tiff_datetime_tag_value or datetimeoriginal_tag_value < earliest_tag_value:
                            earliest_tag_value = datetimeoriginal_tag_value
                    elif tag_name == 'DateTimeDigitized' and value != '':
                        # some bug with midnight hours  00:15:07 appear like 24:15;07, so ' 24:' is replace by ' 00:'
                        str_value=str(value).strip(' ').replace(' 24:', ' 00:')
                        if len(str_value)>19:
                            str_value=str_value[0:18]
                        datetimedigitized_tag_value = datetime.strptime(str_value, '%Y:%m:%d %H:%M:%S')
                        if earliest_tag_value == default_tiff_datetime_tag_value or datetimedigitized_tag_value < earliest_tag_value:
                            earliest_tag_value = datetimedigitized_tag_value
                except ValueError:
                    print("get_tiff_datetime_datetimeoriginal_datetimedigitized('" + file_path + "') : ValueError datetime.strptime(tag " + tag_name + "='" + str_value + "', '%Y:%m:%d %H:%M:%S')")
            return earliest_tag_value, datetime_tag_value, datetimeoriginal_tag_value, datetimedigitized_tag_value
        except IOError:
            # file_path not an image file
            # print("IOError in get_tiff_datetime_datetimeoriginal_datetimedigitized('"+file_path+"'")
            return earliest_tag_value, datetime_tag_value, datetimeoriginal_tag_value, datetimedigitized_tag_value
        except AttributeError:
            # print(file_path + " : AttributeError")
            # print("AttributeError in get_tiff_datetime_datetimeoriginal_datetimedigitized('"+file_path+"'")
            return earliest_tag_value, datetime_tag_value, datetimeoriginal_tag_value, datetimedigitized_tag_value


def get_tiff_datetime_tag_value(file_path):
    # retrieve TIFF 'DateTime' tag value of file_path and convert as date_time object
    date_time_str = get_tiff_tag_value(file_path, 'DateTime')
    if date_time_str != "" and date_time_str != "TIFF tag DateTime Not Found":
        date_time_object = datetime.strptime(date_time_str, '%Y:%m:%d %H:%M:%S')
    else:
        date_time_object = datetime(1970, 1, 1)
    return date_time_object


def get_tiff_tag_value(file_path, tag_name):
    if is_image_file(file_path):
        try:
            for (tag, value) in Image.open(file_path)._getexif().items():
                if str(TAGS.get(tag)) == tag_name:
                    return str(value)
            return "TIFF tag " + tag_name + " Not Found"
        except IOError:
            # file_path not an image file
            return ""
        except AttributeError:
            # print(file_path + " : AttributeError")
            return ""
    return ""


def set_st_mtime_status_value_with_TIFF_DateTime_tag_value_of_all_images(root_folder_path, output_excel_file_path, output_csv_file_path) :
    print('find_duplicate_filename_in_tree_structure(' + root_folder_path + ')\n')
    if os.path.isdir(root_folder_path) :
        
        o_folder_walk=os.walk(root_folder_path)
        d_folder_walk=o_folder_walk

        nb_files=0
        files_list = []
        stdout.write("\rAnalysis of '{0}' : in progress...".format(root_folder_path))
        stdout.flush()
        for (d_folder, d_dirs, d_files) in d_folder_walk :
            for d_file in d_files:
                if 'synoreport' not in d_folder and '#recycle' d_folder not in and is_image_file(d_file):
                    nb_files = nb_files + 1
                    # append full_file_path
                    files_list.append(os.path.join(d_folder, d_file))
        stdout.write("\rAnalysis of '{0}' : {1} files found, retrieving files with bad  st_mtime value in progress...\n".format(root_folder_path, nb_files))
        
        files_list_with_bad_date = []
        step_progress=int(round(nb_files/1000))
        if step_progress == 0:
            step_progress = 1
        prev_nb_files = 0
        curr_nb_files = 0
        nb_files_with_bad_st_mtime=0
        #update progress bar to 0%
        sys_stdout_progress_bar(curr_nb_files, nb_files,"files")
        for d_file_path in files_list :
                curr_nb_files = curr_nb_files + 1
                
                #cmp st_mtime status date value and TIFF 'DateTime' tag value
                file_last_modified = datetime.fromtimestamp(os.stat(d_file_path).st_mtime)
                earliest_tiff_datetime = get_tiff_datetime_datetimeoriginal_datetimedigitized(d_file_path)[0]
                if earliest_tiff_datetime != default_tiff_datetime_tag_value and earliest_tiff_datetime < file_last_modified:
                    nb_files_with_bad_st_mtime = nb_files_with_bad_st_mtime + 1
                    files_list_with_bad_date.append((d_file_path, file_last_modified,earliest_tiff_datetime))
                    sys_stdout_progress_bar(curr_nb_files, nb_files,"files"," : " + str(nb_files_with_bad_st_mtime) + " files with bad st_mtime status date value")
                if ( prev_nb_files + step_progress ) <= curr_nb_files : 
                    prev_nb_files = curr_nb_files
                    sys_stdout_progress_bar(curr_nb_files, nb_files,"files"," : " + str(nb_files_with_bad_st_mtime) + " files with bad st_mtime status date value")
        #last update progress bar to reach 100%
        sys_stdout_progress_bar(curr_nb_files, nb_files,"files")
        stdout.write("\n")
        
        if nb_files_with_bad_st_mtime > 0:
            stdout.write("\r" + str(nb_files_with_bad_st_mtime) + " files found with bad st_mtime os status value, updating in progress...\n")
        
            first_line="file path;st_mtime before;eraliest TIFF DateTime Tag value;st_mtime after"
            split_first_line=first_line.split(';')
            nb_columns=len(split_first_line)
            file_path_column=0
            if output_excel_file_path != '':
                wb = Workbook()
                ws=wb.active
                for j in range(nb_columns):
                    ws.cell(1,j+1).value = split_first_line[j]
                    if split_first_line[j] == "file path":
                        file_path_column=j+1
                is_excel_output=True
            else:
                is_excel_output=False
            
            if output_csv_file_path != '' :
                output_csv_file = open(output_csv_file_path,'w')
                output_csv_file.write(first_line + "\n")
                is_csv_output=True
            else:
                is_csv_output=False
            
            
            
            
            
            #Update the access and modified times of the all files 
            step_progress=int(round(nb_files_with_bad_st_mtime/1000))
            if step_progress == 0:
                step_progress = 1
            prev_nb_files = 0
            i = 0
            #update progress bar to 0%
            sys_stdout_progress_bar(0, nb_files_with_bad_st_mtime,"files")
            for d_file_path, file_last_modified, earliest_tiff_datetime in files_list_with_bad_date :
                i = i + 1
                #sets the access and modified times of the file specified by path.
                try:
                    os.utime(d_file_path, (earliest_tiff_datetime.timestamp(), earliest_tiff_datetime.timestamp()))
                except OSError:
                    print("OSError: os.utime(" + d_file_path + ", (" + str(earliest_tiff_datetime.timestamp()) + ", " + str(earliest_tiff_datetime.timestamp())+"))")
                new_file_last_modified=datetime.fromtimestamp(os.stat(d_file_path).st_mtime)
                if is_excel_output:
                    ws.cell(1 + i, 1).value = d_file_path
                    ws.cell(1 + i, 1).hyperlink = d_file_path
                    ws.cell(1 + i, 1).style = "Hyperlink"
                    ws.cell(1 + i, 2).value = file_last_modified
                    ws.cell(1 + i, 3).value = earliest_tiff_datetime
                    ws.cell(1 + i, 4).value = new_file_last_modified
                
                if is_csv_output:
                    output_csv_file.write(d_file_path + ";" + str(file_last_modified) + ";" + str(earliest_tiff_datetime) + ";" + str(new_file_last_modified) + '\n')
                
                if ( prev_nb_files + step_progress ) <= i : 
                    prev_nb_files = i
                    sys_stdout_progress_bar(i, nb_files_with_bad_st_mtime,"files")
            #last update progress bar to reach 100%
            sys_stdout_progress_bar(i, nb_files_with_bad_st_mtime,"files")
            stdout.write("\n")
            if is_excel_output:
                wb.save(output_excel_file_path)
                stdout.write("File '" + output_excel_file_path + "' Generated\n")
            if is_csv_output:
                output_csv_file.close()
                stdout.write("File '" + output_csv_file_path + "' Generated\n")
        else:
            stdout.write("\r0 file found with bad st_mtime os status value\n")

def main() :
    
    #ArgumentParser(prog=None, usage=None, description=None, epilog=None, parents=[], formatter_class=argparse.HelpFormatter, prefix_chars='-', fromfile_prefix_chars=None, argument_default=None, conflict_handler='error', add_help=True, allow_abbrev=True, exit_on_error=True)
    parser = argparse.ArgumentParser()
    
    
    parser.add_argument('-f', '--folder_path', help='specify the path of the folder in which to find duplicate files. Default is current dir')
    parser.add_argument('-e', '--excel_output', help='specify the Excel output .xlsx file path')
    parser.add_argument('-c', '--csv_output', help='specify the CSV output .csv file path')
    args = parser.parse_args()
    
    
    if ( args.folder_path != '' and args.folder_path is not None ) and ( ( args.csv_output != '' and args.csv_output is not None ) or  ( args.excel_output != '' and args.excel_output is not None ) ):
        if ( args.excel_output != '' and args.excel_output is not None ):
            args.excel_output = os.path.realpath(args.excel_output)
        else:
            args.excel_output = ''
            
        if ( args.csv_output != '' and args.csv_output is not None ):
            args.csv_output = os.path.realpath(args.csv_output)
        else:
            args.csv_output = ''
        
        set_st_mtime_status_value_with_TIFF_DateTime_tag_value_of_all_images(os.path.realpath(args.folder_path), args.excel_output, args.csv_output)
    else :
        parser.print_help()


if __name__ == "__main__":
    # execute only if run as a script
    main()
